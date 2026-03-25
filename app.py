import streamlit as st
import face_recognition
import cv2
import os
import numpy as np
from datetime import datetime
from openpyxl import Workbook, load_workbook
import pandas as pd

# ------------------ PAGE CONFIG ------------------
st.set_page_config(page_title="AI Attendance System", layout="wide")

st.title("📷 AI Smart Attendance System")
st.markdown("### Face Recognition Based Attendance")

# ------------------ LOAD FACES ------------------
path = "faces"
images = []
classNames = []

for img_name in os.listdir(path):
    img_path = os.path.join(path, img_name)
    img = face_recognition.load_image_file(img_path)
    images.append(img)
    classNames.append(os.path.splitext(img_name)[0])

def findEncodings(images):
    encodeList = []
    for img in images:
        encodes = face_recognition.face_encodings(img)
        if len(encodes) > 0:
            encodeList.append(encodes[0])
    return encodeList

encodeListKnown = findEncodings(images)

# ------------------ ATTENDANCE ------------------
def markAttendance(name):
    file_name = "Attendance.xlsx"
    today = datetime.now().strftime("%Y-%m-%d")
    time_now = datetime.now().strftime("%H:%M:%S")

    if not os.path.isfile(file_name):
        wb = Workbook()
        ws = wb.active
        ws.title = today
        ws.append(["Name", "Time", "Status"])
        wb.save(file_name)

    wb = load_workbook(file_name)

    if today not in wb.sheetnames:
        wb.create_sheet(title=today)
        ws = wb[today]
        ws.append(["Name", "Time", "Status"])
    else:
        ws = wb[today]

    names = [row[0].value for row in ws.iter_rows(min_row=2)]

    if name not in names:
        ws.append([name, time_now, "Present"])
        wb.save(file_name)
        return "Present"
    else:
        return "Already Marked"

# ------------------ CAMERA INPUT ------------------
st.sidebar.header("⚙️ Controls")

img_file = st.camera_input("📸 Capture Image")

attendance_list = []

if img_file is not None:
    # Convert image to array
    file_bytes = np.asarray(bytearray(img_file.read()), dtype=np.uint8)
    img = cv2.imdecode(file_bytes, 1)

    imgRGB = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)

    facesCurFrame = face_recognition.face_locations(imgRGB)
    encodesCurFrame = face_recognition.face_encodings(imgRGB, facesCurFrame)

    if len(encodesCurFrame) == 0:
        st.error("❌ No Face Detected")
    else:
        for encodeFace, faceLoc in zip(encodesCurFrame, facesCurFrame):
            matches = face_recognition.compare_faces(encodeListKnown, encodeFace)
            faceDis = face_recognition.face_distance(encodeListKnown, encodeFace)
            matchIndex = np.argmin(faceDis)

            if matches[matchIndex]:
                name = classNames[matchIndex].upper()
                status = markAttendance(name)
            else:
                name = "Unknown"
                status = "Not Registered"

            attendance_list.append([name, datetime.now().strftime("%H:%M:%S"), status])

            # Draw rectangle
            y1, x2, y2, x1 = faceLoc
            cv2.rectangle(img, (x1, y1), (x2, y2), (0,255,0), 2)
            cv2.putText(img, name, (x1, y1-10),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0,255,0), 2)

    st.image(img, channels="BGR")

    df = pd.DataFrame(attendance_list, columns=["Name", "Time", "Status"])
    st.success("✅ Attendance Processed")
    st.dataframe(df)
